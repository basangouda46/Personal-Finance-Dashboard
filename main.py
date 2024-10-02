# Version 2.1 - 01/02/2024 - fixed balance sheet folder bug
# Version 2.2 - 01/03/2024 - save amount as int rather than text


from customtkinter import *
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from datetime import date
import os
import tkinter
from tkinter import messagebox
import time
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, 
NavigationToolbar2Tk)
import pandas as pd

def update_graph():

    global filepath, df, income_df, asset_df, expenses_df, liability_df

    df = pd.read_excel(filepath)

    income_df = df[["Unnamed: 1", "Unnamed: 2", "Unnamed: 3"]]
    income_df = income_df.rename({"Unnamed: 1":"DATE", "Unnamed: 2":"AMOUNT", "Unnamed: 3":"TYPE" }, axis="columns")
    income_df = income_df.drop([0,1])
    income_df['DATE'] = pd.to_datetime(income_df['DATE']).dt.strftime("%m/%d/%y")

    expenses_df = df[["Unnamed: 5", "Unnamed: 6", "Unnamed: 7"]]
    expenses_df = expenses_df.rename({"Unnamed: 5":"DATE", "Unnamed: 6":"AMOUNT", "Unnamed: 7":"TYPE" }, axis="columns")
    expenses_df = expenses_df.drop([0,1])
    expenses_df['DATE'] = pd.to_datetime(expenses_df['DATE']).dt.strftime("%m/%d/%y")

    asset_df = df[["Unnamed: 9", "Unnamed: 10", "Unnamed: 11"]]
    asset_df = asset_df.rename({"Unnamed: 9":"DATE", "Unnamed: 10":"AMOUNT", "Unnamed: 11":"TYPE" }, axis="columns")
    asset_df = asset_df.drop([0,1])
    asset_df['DATE'] = pd.to_datetime(asset_df['DATE']).dt.strftime("%m/%d/%y")

    liability_df = df[["Unnamed: 13", "Unnamed: 14", "Unnamed: 15"]]
    liability_df = liability_df.rename({"Unnamed: 13":"DATE", "Unnamed: 14":"AMOUNT", "Unnamed: 15":"TYPE" }, axis="columns")
    liability_df = liability_df.drop([0,1])
    liability_df['DATE'] = pd.to_datetime(liability_df['DATE']).dt.strftime("%m/%d/%y")

    expenses_df['MONTH'] = pd.to_datetime(expenses_df['DATE']).dt.month

    plot_barchart()
    plot_piechart()


def plot_barchart():
    x = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    #should be 0 every iteration
    y = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    global expenses_df

    for i in expenses_df.index:
        y[(int(expenses_df['MONTH'][i])) - 1] += int(expenses_df['AMOUNT'][i])
        
    #print(y)

    fig = plt.figure(figsize=(6, 2))
    plt.bar(x=x, height=y)

    # You can make your x axis labels vertical using the rotation

    plt.xticks(x)

    # specify the window as master
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().grid(row=0, column=1, rowspan=2, padx=(50,50))#, ipadx=40, ipady=20)

def plot_piechart():

    y = [0, 0, 0, 0]
    labels = ['Assets', 'Liabilities', 'Expenses', 'Misc']

    global income_df, asset_df, liability_df, expenses_df

    income_df = income_df.fillna(0)

    total_income = income_df['AMOUNT'].astype(int).sum()

    expenses_df = expenses_df.fillna(0)
    liability_df = liability_df.fillna(0)
    asset_df = asset_df.fillna(0)

    total_asset = asset_df['AMOUNT'].astype(int).sum()

    total_liability = liability_df['AMOUNT'].astype(int).sum()

    total_expenses = expenses_df['AMOUNT'].astype(int).sum()

    y[0] = (total_asset/total_income) * 100

    y[1] = (total_liability/total_income) * 100

    y[2] = (total_expenses/total_income) * 100

    y[3] = ((total_income - (total_asset + total_liability + total_expenses))/total_income) * 100

    print(y)

    try:
        fig2 = plt.figure(figsize=(4, 2))
        plt.pie(y, labels=labels, autopct='%.0f%%')
    except: 
        y = [30, 20, 47, 3]
        fig2 = plt.figure(figsize=(4, 2))
        plt.pie(y, labels=labels, autopct='%.0f%%')
        print("pie chart exception")


    # specify the window as master
    canvas = FigureCanvasTkAgg(fig2, master=root)
    canvas.draw()
    canvas.get_tk_widget().grid(row=2, column=1, rowspan =2)#, ipadx=40, ipady=20)


 
def income_button_press():

    date_column = xl_sheet['B']

    #max_row_b and date column indexes are off by 1 
    max_row_b = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_b += 1


    c1 = xl_sheet.cell(row=max_row_b, column=2)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_b, column=3)
    income_value = int(income_entry.get())
    if(income_value != 0):
        c2.value = income_value
    income_entry.delete(0, "end")

    c3 =  xl_sheet.cell(row=max_row_b, column=4)
    c3.value  = income_dropdown.get()

    xl_file.save(filename=filepath)

    update_graph()

def expenses_button_press():

    date_column = xl_sheet['F']
    # max_row_f = len(date_column) + 1

    max_row_f = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_f += 1

    c1 = xl_sheet.cell(row=max_row_f, column=6)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_f, column=7)
    expense_value = int(expense_entry.get())
    if(expense_value != 0):
        c2.value = expense_value
    expense_entry.delete(0, "end")

    c3 =  xl_sheet.cell(row=max_row_f, column=8)
    c3.value  = expense_dropdown.get()

    xl_file.save(filename=filepath)

    update_graph()

def asset_button_press():

    date_column = xl_sheet['J']
    # max_row_j = len(date_column) + 1

    max_row_j = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_j += 1

    c1 = xl_sheet.cell(row=max_row_j, column=10)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_j, column=11)
    asset_value = int(asset_entry.get())
    if(asset_value != 0):
        c2.value = asset_value
    asset_entry.delete(0, "end")

    c3 =  xl_sheet.cell(row=max_row_j, column=12)
    c3.value  = asset_dropdown.get()

    xl_file.save(filename=filepath)

    update_graph()

def liability_button_press():

    date_column = xl_sheet['N']
    # max_row_n = len(date_column) + 1

    max_row_n = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_n += 1

    c1 = xl_sheet.cell(row=max_row_n, column=14)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_n, column=15)
    liability_value = int(liability_entry.get())
    if(liability_value != 0):
        c2.value = liability_value 
    liability_entry.delete(0, "end")

    c3 =  xl_sheet.cell(row=max_row_n, column=16)
    c3.value  = liability_dropdown.get()

    xl_file.save(filename=filepath)

    update_graph()

def popup_button_press():
    global filepath_dir, lock
    filepath_dir = tkinter.filedialog.askdirectory()
    popup.destroy()
    popup.update()
    filepath_txt = open("filepath.txt", 'w')
    filepath_txt.write("yes\n")
    filepath_txt.write(filepath_dir + "\n")
    filepath_txt.close()
    lock.set(True)


print("running main.py")

# DISPLAY UI ----------------------------------------------------------------------------------
set_appearance_mode("light")
set_default_color_theme("blue")


root = CTk()
root.title("Personal Finance Dashboard")

date_value = date.today().strftime("%m/%d/%y")

income_types = ["Salary",
                "Loan",
                "Assets",
                "Splitwise",
                "Tax Rebate",
                "Gift"]

expense_types = ["Rent",
                 "Credit Card",
                 "Splitwise"]

asset_types = ["Stocks",
               "Gold",
               "Land",
               "Savings"]

liability_types = ["Student Loan",
                   "Car Loan",
                   "Home Loan"]


window_income = CTkFrame(root, width=500, height=500)
window_income.grid(row=0, column = 0, padx=10, pady=5)

window_expense = CTkFrame(root, width=500, height=500)
window_expense.grid(row=1, column = 0, padx=10, pady=5)

window_asset = CTkFrame(root, width=500, height=500)
window_asset.grid(row=2, column = 0, padx=10, pady=5)

window_liabilities = CTkFrame(root, width=500, height=500)
window_liabilities.grid(row=3, column = 0, padx=10, pady=5)


CTkLabel(window_income, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_income, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_income, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
income_entry = CTkEntry(window_income)
income_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_income, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
income_dropdown = CTkOptionMenu(window_income, values=income_types)
income_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_income, text="Add Income", command=income_button_press).grid(row=7, column=2, padx=5, pady=5)

CTkLabel(window_expense, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_expense, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_expense, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
expense_entry = CTkEntry(window_expense)
expense_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_expense, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
expense_dropdown = CTkOptionMenu(window_expense, values=expense_types)
expense_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_expense, text="Add Expense", command=expenses_button_press).grid(row=7, column=2, padx=5, pady=5)

CTkLabel(window_asset, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_asset, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_asset, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
asset_entry = CTkEntry(window_asset)
asset_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_asset, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
asset_dropdown = CTkOptionMenu(window_asset, values=asset_types)
asset_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_asset, text="Add Asset", command=asset_button_press).grid(row=7, column=2, padx=5, pady=5)

CTkLabel(window_liabilities, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_liabilities, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_liabilities, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
liability_entry = CTkEntry(window_liabilities)
liability_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_liabilities, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
liability_dropdown = CTkOptionMenu(window_liabilities, values=liability_types)
liability_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_liabilities, text="Add Liability", command=liability_button_press).grid(row=7, column=2, padx=5, pady=5)

#-------------------------------------------------------------------------------------------------

# check if file path is provided -----------------------------------------------------------------

filepath_dir = ""

lock = tkinter.BooleanVar(value=False)

filepath_txt = open("filepath.txt", 'r')

data = filepath_txt.readlines()

filepath_txt.close()

if data[0].rstrip() == "no":
    popup = CTkToplevel()
    popup.title("Dialog")
    popup_label = CTkLabel(popup, text="Please choose where you want to save the Balance Sheet")
    popup_label.pack(padx=20, pady=20)
    popup_button = CTkButton(popup, text="OK", command=popup_button_press)
    popup_button.pack(padx=100, pady=20)
    #wait for user to select file location
    root.wait_variable(lock)
elif data[0].rstrip() == "yes":
    filepath_dir = data[1].rstrip()


# ------------------------------------------------------------------------------------------------

# Create new file and create template if file does not exist -------------------------------------
# Create dataframe from the csv file -------------------------------------------------------------

#create filename string
filename = "Balance Sheet - " + date.today().strftime("%Y") + ".xlsx"

filepath = os.path.join(filepath_dir, "Balance Sheet", filename)
# print(filepath)

#handle if file not found and create new file
try:
    xl_file = openpyxl.load_workbook(filepath)
    print("file found")
except:
    print("file not found")
    xl_file = openpyxl.Workbook() 
    xl_sheet= xl_file.active

    yellow = "00FFFF00"
    green = "90EE90"

    #income
    income_heading = xl_sheet.cell(row=2, column=2)
    income_heading.value = "INCOME"
    xl_sheet.merge_cells(start_row=2,start_column=2,end_row=2,end_column=4)
    income_heading.alignment = Alignment(horizontal = 'center')
    income_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    income_date_heading = xl_sheet.cell(row=3, column=2)
    income_date_heading.value = "DATE"
    income_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    income_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    income_amount_heading = xl_sheet.cell(row=3, column=3)
    income_amount_heading.value = "AMOUNT"
    income_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    income_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    income_type_heading = xl_sheet.cell(row=3,  column=4)
    income_type_heading.value = "TYPE"
    income_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    income_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    #expenses
    expense_heading = xl_sheet.cell(row=2, column=6)
    expense_heading.value = "EXPENSES"
    xl_sheet.merge_cells(start_row=2,start_column=6,end_row=2,end_column=8)
    expense_heading.alignment = Alignment(horizontal = 'center')
    expense_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    expense_date_heading = xl_sheet.cell(row=3, column=6)
    expense_date_heading.value = "DATE"
    expense_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    expense_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    expense_amount_heading = xl_sheet.cell(row=3, column=7)
    expense_amount_heading.value = "AMOUNT"
    expense_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    expense_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    expense_type_heading = xl_sheet.cell(row=3,  column=8)
    expense_type_heading.value = "TYPE"
    expense_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    expense_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    #asset
    asset_heading = xl_sheet.cell(row=2, column=10)
    asset_heading.value = "ASSET"
    xl_sheet.merge_cells(start_row=2,start_column=10,end_row=2,end_column=12)
    asset_heading.alignment = Alignment(horizontal = 'center')
    asset_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    asset_date_heading = xl_sheet.cell(row=3, column=10)
    asset_date_heading.value = "DATE"
    asset_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    asset_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    asset_amount_heading = xl_sheet.cell(row=3, column=11)
    asset_amount_heading.value = "AMOUNT"
    asset_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    asset_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    asset_type_heading = xl_sheet.cell(row=3,  column=12)
    asset_type_heading.value = "TYPE"
    asset_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    asset_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    #liability
    liability_heading = xl_sheet.cell(row=2, column=14)
    liability_heading.value = "LIABILITY"
    xl_sheet.merge_cells(start_row=2,start_column=14,end_row=2,end_column=16)
    liability_heading.alignment = Alignment(horizontal = 'center')
    liability_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    liability_date_heading = xl_sheet.cell(row=3, column=14)
    liability_date_heading.value = "DATE"
    liability_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    liability_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    liability_amount_heading = xl_sheet.cell(row=3, column=15)
    liability_amount_heading.value = "AMOUNT"
    liability_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    liability_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    liability_type_heading = xl_sheet.cell(row=3,  column=16)
    liability_type_heading.value = "TYPE"
    liability_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    liability_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    try:
        os.mkdir(os.path.join(filepath_dir, "Balance Sheet"))
    except:
        print("directory already created")

    xl_file.save(filename = filepath)

#PREPARING DATASET-------------------------------------------------------------------------------------

df = pd.read_excel(filepath)

df.head(5)

income_df = df[["Unnamed: 1", "Unnamed: 2", "Unnamed: 3"]]
income_df = income_df.rename({"Unnamed: 1":"DATE", "Unnamed: 2":"AMOUNT", "Unnamed: 3":"TYPE" }, axis="columns")
income_df = income_df.drop([0,1])
income_df['DATE'] = pd.to_datetime(income_df['DATE']).dt.strftime("%m/%d/%y")

expenses_df = df[["Unnamed: 5", "Unnamed: 6", "Unnamed: 7"]]
expenses_df = expenses_df.rename({"Unnamed: 5":"DATE", "Unnamed: 6":"AMOUNT", "Unnamed: 7":"TYPE" }, axis="columns")
expenses_df = expenses_df.drop([0,1])
expenses_df['DATE'] = pd.to_datetime(expenses_df['DATE']).dt.strftime("%m/%d/%y")

asset_df = df[["Unnamed: 9", "Unnamed: 10", "Unnamed: 11"]]
asset_df = asset_df.rename({"Unnamed: 9":"DATE", "Unnamed: 10":"AMOUNT", "Unnamed: 11":"TYPE" }, axis="columns")
asset_df = asset_df.drop([0,1])
asset_df['DATE'] = pd.to_datetime(asset_df['DATE']).dt.strftime("%m/%d/%y")

liability_df = df[["Unnamed: 13", "Unnamed: 14", "Unnamed: 15"]]
liability_df = liability_df.rename({"Unnamed: 13":"DATE", "Unnamed: 14":"AMOUNT", "Unnamed: 15":"TYPE" }, axis="columns")
liability_df = liability_df.drop([0,1])
liability_df['DATE'] = pd.to_datetime(liability_df['DATE']).dt.strftime("%m/%d/%y")

expenses_df['MONTH'] = pd.to_datetime(expenses_df['DATE']).dt.month

#DISPLAY GRAPH----------------------------------------------------------------------------------------

plot_barchart()
plot_piechart()

#-----------------------------------------------------------------------------------------------------

xl_sheet = xl_file.active

root.mainloop()
