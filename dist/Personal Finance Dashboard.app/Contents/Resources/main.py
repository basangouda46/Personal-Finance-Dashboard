from customtkinter import *
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from datetime import date
import os
import tkinter
from tkinter import messagebox
import time



def income_button_press():

    date_column = xl_sheet['B']

    #max_row_b and date column indexes are off by 1 
    max_row_b = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_b += 1


    c1 = xl_sheet.cell(row=max_row_b, column=2)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_b, column=3)
    c2.value = income_entry.get()

    c3 =  xl_sheet.cell(row=max_row_b, column=4)
    c3.value  = income_dropdown.get()

    xl_file.save(filename=filepath)

def expenses_button_press():

    date_column = xl_sheet['F']
    # max_row_f = len(date_column) + 1

    max_row_f = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_f += 1

    c1 = xl_sheet.cell(row=max_row_f, column=6)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_f, column=7)
    c2.value = expense_entry.get()

    c3 =  xl_sheet.cell(row=max_row_f, column=8)
    c3.value  = expense_dropdown.get()

    xl_file.save(filename=filepath)

def asset_button_press():

    date_column = xl_sheet['J']
    # max_row_j = len(date_column) + 1

    max_row_j = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_j += 1

    c1 = xl_sheet.cell(row=max_row_j, column=10)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_j, column=11)
    c2.value = asset_entry.get()

    c3 =  xl_sheet.cell(row=max_row_j, column=12)
    c3.value  = asset_dropdown.get()

    xl_file.save(filename=filepath)

def liability_button_press():

    date_column = xl_sheet['N']
    # max_row_n = len(date_column) + 1

    max_row_n = 2
    date_column = date_column[1:]
    for r in date_column:
        if r.value is None:
            break
        else:
            max_row_n += 1

    c1 = xl_sheet.cell(row=max_row_n, column=14)
    c1.value = date_value

    c2 =  xl_sheet.cell(row=max_row_n, column=15)
    c2.value = liability_entry.get()

    c3 =  xl_sheet.cell(row=max_row_n, column=16)
    c3.value  = liability_dropdown.get()

    xl_file.save(filename=filepath)

def popup_button_press():
    global filepath_dir, lock
    filepath_dir = tkinter.filedialog.askdirectory()
    popup.destroy()
    popup.update()
    filepath_txt = open("filepath.txt", 'w')
    filepath_txt.write("yes\n")
    filepath_txt.write(filepath_dir + "\n")
    filepath_txt.close()
    lock.set(True)


print("running main.py")

# DISPLAY UI ----------------------------------------------------------------------------------
set_appearance_mode("light")
set_default_color_theme("blue")


root = CTk()
root.title("Personal Finance Dashboard")

date_value = date.today().strftime("%m/%d/%y")

income_types = ["Salary",
                "Loan",
                "Assets",
                "Splitwise",
                "Tax Rebate",
                "Gift"]

expense_types = ["Rent",
                 "Credit Card",
                 "Splitwise"]

asset_types = ["Stocks",
               "Gold",
               "Land"]

liability_types = ["Student Loan",
                   "Car Loan",
                   "Home Loan"]


window_income = CTkFrame(root, width=500, height=500)
window_income.grid(padx=10, pady=5)

window_expense = CTkFrame(root, width=500, height=500)
window_expense.grid(padx=10, pady=5)

window_asset = CTkFrame(root, width=500, height=500)
window_asset.grid(padx=10, pady=5)

window_liabilities = CTkFrame(root, width=500, height=500)
window_liabilities.grid(padx=10, pady=5)


CTkLabel(window_income, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_income, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_income, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
income_entry = CTkEntry(window_income)
income_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_income, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
income_dropdown = CTkOptionMenu(window_income, values=income_types)
income_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_income, text="Add Income", command=income_button_press).grid(row=7, column=2, padx=5, pady=5)

CTkLabel(window_expense, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_expense, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_expense, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
expense_entry = CTkEntry(window_expense)
expense_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_expense, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
expense_dropdown = CTkOptionMenu(window_expense, values=expense_types)
expense_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_expense, text="Add Expense", command=expenses_button_press).grid(row=7, column=2, padx=5, pady=5)

CTkLabel(window_asset, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_asset, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_asset, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
asset_entry = CTkEntry(window_asset)
asset_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_asset, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
asset_dropdown = CTkOptionMenu(window_asset, values=asset_types)
asset_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_asset, text="Add Asset", command=asset_button_press).grid(row=7, column=2, padx=5, pady=5)

CTkLabel(window_liabilities, text="Date", justify="left").grid(sticky=W, row=3, column=0, padx=5, pady=5)
CTkLabel(window_liabilities, text=date_value).grid(row=3, column=1, padx=5, pady=5)
CTkLabel(window_liabilities, text="Amount", justify="left").grid(sticky=W, row=4, column=0, padx=5, pady=5)
liability_entry = CTkEntry(window_liabilities)
liability_entry.grid(row=4, column=1, padx=5, pady=5)
CTkLabel(window_liabilities, text="Type", justify="left").grid(sticky=W, row=5, column=0, padx=5, pady=5)
liability_dropdown = CTkOptionMenu(window_liabilities, values=liability_types)
liability_dropdown.grid(row=5, column=1, padx=5, pady=5)
CTkButton(window_liabilities, text="Add Liability", command=liability_button_press).grid(row=7, column=2, padx=5, pady=5)

#-------------------------------------------------------------------------------------------------

# check if file path is provided -----------------------------------------------------------------

filepath_dir = ""

lock = tkinter.BooleanVar(value=False)

filepath_txt = open("filepath.txt", 'r')

data = filepath_txt.readlines()

filepath_txt.close()

if data[0].rstrip() == "no":
    popup = CTkToplevel()
    popup.title("Dialog")
    popup_label = CTkLabel(popup, text="Please choose where you want to save the Balance Sheet")
    popup_label.pack(padx=20, pady=20)
    popup_button = CTkButton(popup, text="OK", command=popup_button_press)
    popup_button.pack(padx=100, pady=20)
    #wait for user to select file location
    root.wait_variable(lock)
elif data[0].rstrip() == "yes":
    filepath_dir = data[1].rstrip()

# ------------------------------------------------------------------------------------------------



# Create new file and create template if file does not exist -------------------------------------

#create filename string
filename = "Balance Sheet - " + date.today().strftime("%Y") + ".xlsx"

filepath = os.path.join(filepath_dir, "Balance Sheet", filename)
# print(filepath)

#handle if file not found and create new file
try:
    xl_file = openpyxl.load_workbook(filepath)
    print("file found")
except:
    print("file not found")
    xl_file = openpyxl.Workbook() 
    xl_sheet= xl_file.active

    yellow = "00FFFF00"
    green = "90EE90"

    #income
    income_heading = xl_sheet.cell(row=2, column=2)
    income_heading.value = "INCOME"
    xl_sheet.merge_cells(start_row=2,start_column=2,end_row=2,end_column=4)
    income_heading.alignment = Alignment(horizontal = 'center')
    income_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    income_date_heading = xl_sheet.cell(row=3, column=2)
    income_date_heading.value = "DATE"
    income_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    income_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    income_amount_heading = xl_sheet.cell(row=3, column=3)
    income_amount_heading.value = "AMOUNT"
    income_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    income_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    income_type_heading = xl_sheet.cell(row=3,  column=4)
    income_type_heading.value = "TYPE"
    income_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    income_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    #expenses
    expense_heading = xl_sheet.cell(row=2, column=6)
    expense_heading.value = "EXPENSES"
    xl_sheet.merge_cells(start_row=2,start_column=6,end_row=2,end_column=8)
    expense_heading.alignment = Alignment(horizontal = 'center')
    expense_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    expense_date_heading = xl_sheet.cell(row=3, column=6)
    expense_date_heading.value = "DATE"
    expense_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    expense_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    expense_amount_heading = xl_sheet.cell(row=3, column=7)
    expense_amount_heading.value = "AMOUNT"
    expense_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    expense_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    expense_type_heading = xl_sheet.cell(row=3,  column=8)
    expense_type_heading.value = "TYPE"
    expense_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    expense_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    #asset
    asset_heading = xl_sheet.cell(row=2, column=10)
    asset_heading.value = "ASSET"
    xl_sheet.merge_cells(start_row=2,start_column=10,end_row=2,end_column=12)
    asset_heading.alignment = Alignment(horizontal = 'center')
    asset_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    asset_date_heading = xl_sheet.cell(row=3, column=10)
    asset_date_heading.value = "DATE"
    asset_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    asset_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    asset_amount_heading = xl_sheet.cell(row=3, column=11)
    asset_amount_heading.value = "AMOUNT"
    asset_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    asset_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    asset_type_heading = xl_sheet.cell(row=3,  column=12)
    asset_type_heading.value = "TYPE"
    asset_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    asset_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    #liability
    liability_heading = xl_sheet.cell(row=2, column=14)
    liability_heading.value = "LIABILITY"
    xl_sheet.merge_cells(start_row=2,start_column=14,end_row=2,end_column=16)
    liability_heading.alignment = Alignment(horizontal = 'center')
    liability_heading.fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    liability_date_heading = xl_sheet.cell(row=3, column=14)
    liability_date_heading.value = "DATE"
    liability_date_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    liability_date_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    liability_amount_heading = xl_sheet.cell(row=3, column=15)
    liability_amount_heading.value = "AMOUNT"
    liability_amount_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    liability_amount_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    liability_type_heading = xl_sheet.cell(row=3,  column=16)
    liability_type_heading.value = "TYPE"
    liability_type_heading.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    liability_type_heading.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    os.mkdir(os.path.join(filepath_dir, "Balance Sheet"))

    xl_file.save(filename = filepath)

#-----------------------------------------------------------------------------------------------

xl_sheet = xl_file.active

root.mainloop()
